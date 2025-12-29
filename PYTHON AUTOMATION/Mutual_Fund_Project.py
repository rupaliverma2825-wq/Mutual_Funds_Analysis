import pandas as pd
import matplotlib.pyplot as plt
from sqlalchemy import create_engine

engine = create_engine(
    "postgresql+psycopg2://postgres:Rupali0055@localhost:5432/mutual_fund"
)
# TOP 5 FUNDS ON THE BASIS OF ONE MONTH RETURN

one_month_return = pd.read_sql(
    """
    SELECT
        FUND_NAME,
        ONE_MONTH_RETURN
    FROM
        MUTUAL_FUNDS
    ORDER BY
        ONE_MONTH_RETURN DESC
    LIMIT
        5;
    """, engine
)

print(one_month_return)

plt.figure(figsize=(8,6))
plt.bar(one_month_return["fund_name"],one_month_return["one_month_return"])
plt.title("TOP 5 FUNDS ON THE BASIS OF ONE MONTH RETURN")
plt.xticks(rotation=45, ha="right")
plt.tight_layout()
plt.savefig("one_month_return.png")
plt.show()

# FUND NAMES WHERE MORNING STAR RATING AMD VALUE RESEARCH RATING = 5 (1 YEAR RETURN )

m_star= pd.read_sql(
   """ SELECT
	FUND_NAME,
	ONE_YEAR_RETURN,
	MORNING_STAR_RATING,
	VALUE_RESEARCH_RATING
FROM
	MUTUAL_FUNDS
WHERE
	MORNING_STAR_RATING = VALUE_RESEARCH_RATING
	AND MORNING_STAR_RATING = 5
ORDER BY
	ONE_YEAR_RETURN DESC
limit 5;""",engine

)

plt.bar(m_star["fund_name"],m_star["one_year_return"])
plt.title("FUND WHERE MSR AND VRR =5")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("m_star_v_rating=5")
plt.show()

# TOP FUND MANAGERS ONE THE BASIS OF ONE MONTH RETURN

top_fund_manager=pd.read_sql(
"""    SELECT
	FUND_NAME,
	FUND_MANAGER,
	ONE_MONTH_RETURN
FROM
	MUTUAL_FUNDS
ORDER BY
	ONE_MONTH_RETURN DESC
 LIMIT 5;""",engine

)

print(top_fund_manager)

plt.bar(top_fund_manager["fund_manager"],top_fund_manager["one_month_return"])
plt.title("TOP FUND MANAGERS = 1M RETURN")
plt.xticks(rotation=45, ha="right")
plt.tight_layout()
plt.savefig("top_fund_manager_1M.png")
plt.show()

# TOP FUND MANAGERS ON THE BASIS OF 3 YEAR RETURNS

top_fund_manager3=pd.read_sql(
    """SELECT
	FUND_NAME,
	FUND_MANAGER,
	THREE_YEAR_RETURN
FROM
	MUTUAL_FUNDS
ORDER BY
	THREE_YEAR_RETURN DESC
LIMIT 5;""",engine
)

plt.figure(figsize=(10,6))
plt.bar(top_fund_manager3["fund_manager"],top_fund_manager3["three_year_return"])
plt.title("TOP FUND MANAGERS = 3Y RETURNS")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("Top_fund_manager_3.png")
plt.show()

# TOP 5 FUNDS WITH THEIR RETURN WHERE RISK= VERY HIGH (1 YEAR RETURN)

very_high_risk= pd.read_sql(
"""    SELECT
	FUND_NAME,
	ONE_YEAR_RETURN,
	RISK
FROM
	MUTUAL_FUNDS
WHERE
	RISK='Very High'
ORDER BY
	ONE_YEAR_RETURN DESC
limit 5;""",engine
)

plt.figure(figsize=(10,6))
plt.bar(very_high_risk["fund_name"],very_high_risk["one_year_return"])
plt.title("TOP 5 FUNDSWHERE RISK= VERY HIGH ,1 YR")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("risk=very high(1 yr).png")
plt.show()

# TOP 5 FUNDS WHERE RISK= VERY HIGH (3 YRS )

very_high_3=pd.read_sql(
	"""SELECT
	FUND_NAME,
	THREE_YEAR_RETURN,
	RISK
FROM
	MUTUAL_FUNDS
WHERE
	RISK = 'Very High'
ORDER BY
	THREE_YEAR_RETURN DESC
LIMIT
	5;""",engine
)

plt.figure(figsize=(10,6))
plt.bar(very_high_3["fund_name"],very_high_3["three_year_return"])
plt.title("TOP 5 FUNDS WHERE RISK= VERY HIGH (3Y)")
plt.xticks(rotation=45,ha="right")
plt.tight_layout()
plt.savefig("risk=very high(3 yr).png")
plt.show()

# COMPARATIVE ANALYSIS OF THREE-YEAR RETURNS ACROSS EQUITY CATEGORIES

equity_3= pd.read_sql(
	"""SELECT
	FUND_NAME,
	THREE_YEAR_RETURN,
	CATEGORY
FROM
	MUTUAL_FUNDS
WHERE
	CATEGORY = 'Equity'
ORDER BY
	THREE_YEAR_RETURN desc
LIMIT
	5;
""",engine
)

plt.figure(figsize=(8,6))
plt.bar(equity_3["fund_name"],equity_3["three_year_return"])
plt.xticks(rotation=45,ha="right")
plt.title("3-YEAR RETURNS (EQUITY CATEGORIES)")
plt.tight_layout()
plt.savefig("equity_3yr")
plt.show()

# COMPARATIVE ANALYSIS OF THREE-YEAR RETURNS ACROSS DEBT CATEGORIES

debt_3yr=pd.read_sql(
	"""SELECT
	FUND_NAME,
	THREE_YEAR_RETURN,
	CATEGORY
FROM
	MUTUAL_FUNDS
WHERE
	CATEGORY = 'Debt'
ORDER BY
	THREE_YEAR_RETURN desc
LIMIT
	5;
""",engine
)

plt.figure(figsize=(8,6))
plt.bar(debt_3yr["fund_name"],debt_3yr["three_year_return"])
plt.xticks(rotation=45,ha="right")
plt.title("3-YEAR RETURNS (DEBT CATEGORIES)")
plt.tight_layout()
plt.savefig("debt_3yr")
plt.show()

# COMPARATIVE ANALYSIS OF THREE-YEAR RETURNS ACROSS HYBRID CATEGORIES

hybrid_3yr=pd.read_sql(
	"""SELECT
	FUND_NAME,
	THREE_YEAR_RETURN,
	CATEGORY
FROM
	MUTUAL_FUNDS
WHERE
	CATEGORY = 'Hybrid'
ORDER BY
	THREE_YEAR_RETURN desc
LIMIT
	5;	
""",engine
)

plt.figure(figsize=(8,6))
plt.bar(hybrid_3yr["fund_name"],hybrid_3yr["three_year_return"])
plt.xticks(rotation=45,ha="right")
plt.title("3-YEAR RETURNS (HYBRID CATEGORIES)")
plt.tight_layout()
plt.savefig("Hybrid_3yr")
plt.show()

# COMPARATIVE ANALYSIS OF THREE-YEAR RETURNS ACROSS OTHER CATEGORIES

others_3yr= pd.read_sql(
	"""SELECT
	FUND_NAME,
	THREE_YEAR_RETURN,
	CATEGORY
FROM
	MUTUAL_FUNDS
WHERE
	CATEGORY = 'Other'
ORDER BY
	THREE_YEAR_RETURN desc
LIMIT
	5;	
""",engine
)

plt.figure(figsize=(8,6))
plt.bar(others_3yr["fund_name"],others_3yr["three_year_return"])
plt.xticks(rotation=45,ha="right")
plt.title("3-YEAR RETURNS (OTHERS CATEGORIES)")
plt.tight_layout()
plt.savefig("others_3yr")
plt.show()

# MOST TRUSTED AMC , FUNDS IN RELATION TO AUM

most_trusted_AMC=pd.read_sql(
	"""SELECT
	AMC,
	FUND_NAME,
	AUM
FROM
	MUTUAL_FUNDS
ORDER BY
	AUM DESC
LIMIT
	5;""",engine
)

plt.figure(figsize=(8,6))
plt.bar(most_trusted_AMC["amc"],most_trusted_AMC["aum"])
plt.xticks(rotation=45,ha="right")
plt.title("MOST TRUSTED AMC BY AUM")
plt.tight_layout()
plt.savefig("Most_trusted_AMC")
plt.show()

# convert it into excel files with chart

from openpyxl import Workbook
from openpyxl.drawing.image import Image

def create_excel_same_sheet(df, chart_path, file_name, insight_text):
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"

# Write table
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(row)

    # Add chart below the table (row after last table row + 2)
    chart_row = len(df) + 3
    ws.add_image(Image(chart_path), f"A{chart_row}")

    # Add insight below chart
    insight_row = chart_row + 20  # leave space for chart
    ws[f"A{insight_row}"] = "Insight:"
    ws[f"B{insight_row}"] = insight_text

    wb.save(file_name)
    print(f"{file_name} created ✅")

create_excel_same_sheet(
    df=one_month_return,                  # Your DataFrame
    chart_path="one_month_return.png",    # Chart file path
    file_name="Top_5_1M_Return.xlsx",    # Excel file name
    insight_text="Kotak Nifty Financial Services Ex Bank Index tops one-month return at 12.49%"
)

files=[
    
[one_month_return,"one_month_return.png","Top_5_1M_Return.xlsx",
"Kotak Nifty Financial Services Ex-Bank Index emerged as the top-performing mutual fund, delivering a 12.49% return." ],
[m_star,"m_star_v_rating=5.png","m_star,_v_rating=5.xlsx","Quant Small Cap–G is a top-performing mutual fund with a 5-star rating from both Morningstar and Value Research, delivering a 54.5% one-year return"],
[top_fund_manager,"top_fund_manager_1M.png","top_fund_manager_1M.xlsx","Abhishek Bisen emerged as a top fund manager based on one-month performance, delivering a 12.49% return"],
[top_fund_manager3,"Top_fund_manager_3.png","top_fund_manager_3Y.xlsx","Kayzad Eghlim emerged as a top fund manager based on three-year performance, delivering a 39.47% return."],
[very_high_risk,"risk=very high(1 yr).png","risk=very high(1 yr).xlsx","HDFC Defence Reg–G is a top-performing fund with very high risk, delivering an 86.35% one-year return."],
[very_high_3,"risk=very high(3 yr).png","risk=very high(3 yr).xlsx","ICICI Pru Bharat 22 FOF–G is a top-performing fund, delivering a 39.47% return over three years."],
[equity_3,"equity_3yr.png","equity_3yr.xlsx","Aditya Birla SL PSU Equity Reg–G is a top-performing fund in the Equity category, delivering a 38.64% return over three years."],
[debt_3yr,"debt_3yr.png","debt_3yr.xlsx","Bank of India Credit Risk Reg–G, a Debt category fund, delivered a 39.07% return over three years"],
[hybrid_3yr,"Hybrid_3yr.png","Hybrid_3_Yr.xlsx","Quant Multi Asset–G is a top-performing fund in the Hybrid category, delivering a 24.84% return."],
[others_3yr,"others_3yr.png","others_3yr.xlsx","ICICI Pru Bharat 22 FOF–G is a top-performing fund in the Other category, delivering a 39.47% return over three years."],
[most_trusted_AMC,"Most_trusted_AMC.png","Most_trusted_AMC.xlsx","HDFC Mutual Fund is the most trusted AMC, managing an AUM of ₹95,391.56 crores."]

]

for i in files:
    create_excel_same_sheet(* i)

